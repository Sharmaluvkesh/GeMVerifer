import io
from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from app.schemas import AnalysisReportResponse

def generate_excel_report(report: AnalysisReportResponse) -> bytes:
    """
    Generates a multi-vendor side-by-side Technical Evaluation Matrix in Excel format (.xlsx).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "GeM Technical Evaluation"
    
    # Styles
    title_font = Font(name="Arial", size=16, bold=True, color="1E293B")
    subtitle_font = Font(name="Arial", size=11, italic=True, color="64748B")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Arial", size=10, bold=True)
    normal_font = Font(name="Arial", size=10)
    
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    tender_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    
    green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")  # COMPLIANT
    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")      # NON_COMPLIANT
    yellow_fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")   # NEEDS_REVIEW
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    # Title Block
    ws['A1'] = "Government e-Marketplace (GeM) - Technical Evaluation Report"
    ws['A1'].font = title_font
    ws['A2'] = f"Item: {report.tender_info.item_name} | Tender ID: {report.tender_info.tender_id or 'N/A'} | Evaluated: {report.evaluated_at[:10]}"
    ws['A2'].font = subtitle_font
    
    ws.append([]) # Empty row
    
    # Summary Row Headers
    summary_headers = ["Parameter / Metric", "Tender Requirement", "Mandatory?"]
    for v in report.vendors:
        summary_headers.extend([f"{v.vendor_name} (Status)", f"{v.vendor_name} (Score: {v.technical_score}%)"])
        
    ws.append(summary_headers)
    header_row_idx = ws.max_row
    
    for col_idx in range(1, len(summary_headers) + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Build matrix of technical parameters
    tender_params = report.tender_info.technical_parameters
    
    for tp in tender_params:
        row_data = [
            tp.parameter_name,
            tp.required_value,
            "YES" if tp.mandatory else "NO"
        ]
        
        for v in report.vendors:
            # find matching result for this vendor
            matched_res = next((r for r in v.parameter_results if r.parameter_name == tp.parameter_name), None)
            if matched_res:
                row_data.extend([matched_res.status, matched_res.vendor_offered_value])
            else:
                row_data.extend(["NOT PROVIDED", "N/A"])
                
        ws.append(row_data)
        current_row = ws.max_row
        
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = normal_font
            cell.border = thin_border
            
            # Apply color coding to status columns
            val_str = str(cell.value)
            if val_str == "COMPLIANT":
                cell.fill = green_fill
                cell.font = bold_font
            elif val_str == "NON_COMPLIANT":
                cell.fill = red_fill
                cell.font = bold_font
            elif val_str == "NEEDS_REVIEW":
                cell.fill = yellow_fill
                cell.font = bold_font

    # Adjust Column Widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_pdf_report(report: AnalysisReportResponse) -> bytes:
    """
    Generates a publication-ready PDF Technical Evaluation Report using ReportLab.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36
    )
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontSize=20,
        leading=24,
        textColor=colors.HexColor('#1E3A8A'),
        fontName='Helvetica-Bold',
        spaceAfter=6
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubTitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#64748B'),
        fontName='Helvetica-Oblique',
        spaceAfter=15
    )
    
    section_heading = ParagraphStyle(
        'SectionHeading',
        parent=styles['Heading2'],
        fontSize=14,
        leading=18,
        textColor=colors.HexColor('#0F172A'),
        fontName='Helvetica-Bold',
        spaceBefore=12,
        spaceAfter=8
    )

    cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontSize=9,
        leading=11,
        textColor=colors.HexColor('#334155')
    )

    cell_header = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontSize=9,
        leading=11,
        textColor=colors.white,
        fontName='Helvetica-Bold'
    )

    story = []
    
    # Title Block
    story.append(Paragraph("GeM Technical Evaluation Report", title_style))
    story.append(Paragraph(f"Tender Item: {report.tender_info.item_name} | Generated: {report.evaluated_at[:10]}", subtitle_style))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#CBD5E1"), spaceAfter=15))
    
    # Vendor Summary Cards Table
    story.append(Paragraph("Executive Summary", section_heading))
    
    summary_data = [
        [Paragraph("Vendor Name", cell_header), Paragraph("Technical Score", cell_header), Paragraph("Disqualification Status", cell_header), Paragraph("Missing Documents", cell_header)]
    ]
    
    for v in report.vendors:
        disq_text = "DISQUALIFIED" if v.is_disqualified else "QUALIFIED"
        disq_color = "#EF4444" if v.is_disqualified else "#10B981"
        disq_para = Paragraph(f"<font color='{disq_color}'><b>{disq_text}</b></font>", cell_style)
        
        missing_text = ", ".join(v.missing_documents) if v.missing_documents else "None"
        
        summary_data.append([
            Paragraph(f"<b>{v.vendor_name}</b>", cell_style),
            Paragraph(f"<b>{v.technical_score}%</b>", cell_style),
            disq_para,
            Paragraph(missing_text, cell_style)
        ])
        
    summary_table = Table(summary_data, colWidths=[140, 90, 130, 160])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A8A')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')])
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 15))
    
    # Detailed Evaluation per Vendor
    for v in report.vendors:
        story.append(Paragraph(f"Detailed Compliance: {v.vendor_name}", section_heading))
        
        param_table_data = [
            [Paragraph("Parameter Name", cell_header), Paragraph("Tender Requirement", cell_header), Paragraph("Offered Value", cell_header), Paragraph("Status", cell_header), Paragraph("Explanation", cell_header)]
        ]
        
        for res in v.parameter_results:
            st = res.status
            st_color = "#10B981" if st == "COMPLIANT" else ("#EF4444" if st == "NON_COMPLIANT" else "#F59E0B")
            st_para = Paragraph(f"<font color='{st_color}'><b>{st}</b></font>", cell_style)
            
            param_table_data.append([
                Paragraph(res.parameter_name, cell_style),
                Paragraph(res.tender_required_value, cell_style),
                Paragraph(res.vendor_offered_value, cell_style),
                st_para,
                Paragraph(res.explanation, cell_style)
            ])
            
        p_table = Table(param_table_data, colWidths=[100, 100, 100, 80, 140])
        p_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#334155')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F1F5F9')])
        ]))
        story.append(p_table)
        story.append(Spacer(1, 15))
        
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()
